"""Build publication-style regression table from district-clustered results."""
from pathlib import Path
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
from src.utils import fmt_coef, fmt_se

POOLED = REPO_ROOT / "outputs" / "tables" / "Regression_Results_Pooled_DistrictCluster.xlsx"
BYSTATE = REPO_ROOT / "outputs" / "tables" / "Regression_Results_By_State.xlsx"
OUT = REPO_ROOT / "outputs" / "tables" / "Regression_Table_Paper.xlsx"


def get_pooled(model):
    coef = pd.read_excel(POOLED, sheet_name=f"{model}_Coef")
    stats = pd.read_excel(POOLED, sheet_name=f"{model}_Stats")
    out = {}
    for var in ["Seasonal_Ratio"]:
        r = coef[coef["Variable"] == var].iloc[0]
        out["flood"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    ndvi_var = "NDVI_median_t_minus_1" if model == "Median" else "NDVI_mean_t_minus_1"
    ndbi_var = "NDBI_median_t_minus_1" if model == "Median" else "NDBI_mean_t_minus_1"
    r = coef[coef["Variable"] == ndvi_var].iloc[0]
    out["ndvi"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    r = coef[coef["Variable"] == ndbi_var].iloc[0]
    out["ndbi"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    obs = int(stats.loc[stats["Item"] == "Observations", "Value"].iloc[0])
    r2 = float(stats.loc[stats["Item"] == "R-sq (within)", "Value"].iloc[0])
    out["obs"] = obs
    out["r2"] = r2
    return out


def get_state(model, state):
    df = pd.read_excel(BYSTATE, sheet_name=f"{model}_All_States")
    out = {}
    coef_col = f"{state}_Coef"
    se_col = f"{state}_SE"
    sig_col = f"{state}_Sig"

    def row(varname):
        r = df[df["Variable"] == varname].iloc[0]
        return (r[coef_col], r[se_col], r[sig_col])

    out["flood"] = row("Seasonal_Ratio")
    ndvi_var = "NDVI_median_t_minus_1" if model == "Median" else "NDVI_mean_t_minus_1"
    ndbi_var = "NDBI_median_t_minus_1" if model == "Median" else "NDBI_mean_t_minus_1"
    out["ndvi"] = row(ndvi_var)
    out["ndbi"] = row(ndbi_var)
    out["obs"] = int(df.loc[df["Variable"] == "Observations", coef_col].iloc[0])
    out["r2"] = float(df.loc[df["Variable"] == "R-sq (within)", coef_col].iloc[0])
    return out


def build_sheet(ws, model_name):
    pooled = get_pooled(model_name)
    bihar = get_state(model_name, "BIHAR")
    jhar = get_state(model_name, "JHARKHAND")
    orissa = get_state(model_name, "ORISSA")
    wb_state = get_state(model_name, "WB")

    cols = [pooled, bihar, jhar, orissa, wb_state]
    headers = ["All states", "Bihar", "Jharkhand", "Odisha", "West Bengal"]

    bold = Font(bold=True)
    italic = Font(italic=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")

    ws.cell(row=1, column=1, value="Dependent variable: Growth in Night Lights")
    ws.cell(row=1, column=1).font = bold
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    header_row = 3
    ws.cell(row=header_row, column=1, value="")
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = bold
        c.alignment = center
        c.border = Border(top=thick, bottom=thin)
    ws.cell(row=header_row, column=1).border = Border(top=thick, bottom=thin)

    var_labels = [
        ("flood", "Proportion of area flooded"),
        ("ndvi", "NDVI_{t-1}"),
        ("ndbi", "NDBI_{t-1}"),
    ]

    r = header_row + 1
    for key, label in var_labels:
        ws.cell(row=r, column=1, value=label).alignment = left
        for j, col_data in enumerate(cols, start=2):
            coef, se, sig = col_data[key]
            ws.cell(row=r, column=j, value=fmt_coef(coef, sig)).alignment = center
        for j, col_data in enumerate(cols, start=2):
            coef, se, sig = col_data[key]
            cell = ws.cell(row=r + 1, column=j, value=fmt_se(se))
            cell.alignment = center
            cell.font = Font(italic=True, size=10)
        r += 2

    fe_top = r
    for label in ["AC FE", "Year FE"]:
        ws.cell(row=r, column=1, value=label).alignment = left
        for j in range(2, 7):
            ws.cell(row=r, column=j, value="Yes").alignment = center
        r += 1

    for j in range(1, 7):
        ws.cell(row=fe_top, column=j).border = Border(top=thin)

    obs_row = r
    ws.cell(row=r, column=1, value="Observations").alignment = left
    for j, col_data in enumerate(cols, start=2):
        ws.cell(row=r, column=j, value=col_data["obs"]).alignment = center
    r += 1

    ws.cell(row=r, column=1, value="R² (within)").alignment = left
    for j, col_data in enumerate(cols, start=2):
        ws.cell(row=r, column=j, value=round(col_data["r2"], 3)).alignment = center
    r += 1

    for j in range(1, 7):
        ws.cell(row=obs_row, column=j).border = Border(top=thin)
        ws.cell(row=r - 1, column=j).border = Border(bottom=thick)

    ws.column_dimensions["A"].width = 32
    for j in range(2, 7):
        ws.column_dimensions[get_column_letter(j)].width = 14


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws_med = wb.create_sheet("Median_Model")
    build_sheet(ws_med, "Median")
    ws_mean = wb.create_sheet("Mean_Model")
    build_sheet(ws_mean, "Mean")

    ws_n = wb.create_sheet("Notes")
    notes = [
        "HOW TO READ THIS TABLE",
        "----------------------",
        "Each variable cell contains TWO numbers stacked vertically:",
        "",
        "    -0.388**       <-- top: estimated coefficient (β̂), with significance stars",
        "    (0.158)        <-- bottom: cluster-robust standard error, in parentheses (italic)",
        "",
        "Significance stars (based on the p-value):",
        "    ***   p < 0.01   (significant at 1% level — very strong evidence effect ≠ 0)",
        "    **    p < 0.05   (significant at 5% level — strong evidence effect ≠ 0)",
        "    *     p < 0.10   (significant at 10% level — weak/marginal evidence)",
        "    (blank)          not statistically distinguishable from 0 at conventional levels",
        "",
        "Standard error (the number in parentheses):",
        "    Measures how precisely the coefficient is estimated. Smaller SE = more precise.",
        "    Here SEs are CLUSTERED BY DISTRICT, which accounts for correlated shocks (e.g. a flood",
        "    hits all ACs in a district together) so we don't overstate statistical confidence.",
        "    Rough rule of thumb: |coefficient / SE| > ~2 means significant at 5% (the ** mark).",
        "",
        "Worked example — West Bengal column, 'Proportion of area flooded' row:",
        "    Coefficient = -0.388, SE = 0.158, stars = ** (p<0.05).",
        "    Interpretation: holding NDVI, NDBI, AC and year fixed, a one-unit rise in the",
        "    seasonal-flood share of an AC is associated with a 0.388 DECREASE in Δlog NL",
        "    (i.e. roughly a 39% slower year-on-year night-light growth) in West Bengal.",
        "    This effect is statistically significant at the 5% level.",
        "",
        "Bottom block (AC FE, Year FE, Observations, R² within):",
        "    These are model-level facts, not estimates with uncertainty, so they appear on a",
        "    single row each — no SE/stars needed.",
        "    AC FE = Yes  -> AC fixed effects included (controls for time-invariant AC features)",
        "    Year FE = Yes-> year fixed effects included (controls for India-wide annual shocks)",
        "    R² (within) -> share of within-AC variation explained by the regressors",
        "                   (after FE absorb level differences). Can be small or negative when",
        "                   the within-AC signal is weak; this is normal in TWFE panels.",
        "",
        "============================================================",
        "",
        "Source files:",
        "  - Regression_Results_Pooled_DistrictCluster.xlsx  (All states column)",
        "  - Regression_Results_By_State.xlsx                (Bihar / Jharkhand / Odisha / WB columns)",
        "",
        "Specification:",
        "  Δlog NL_it = β₀ + β₁·Seasonal_Ratio_it + β₂·NDVI_{i,t-1} + β₃·NDBI_{i,t-1} + α_i + γ_t + ε_it",
        "  α_i = AC fixed effects, γ_t = year fixed effects.",
        "  Estimator: pyfixest.feols (OLS with two-way FE); SEs clustered at district.",
        "",
        "Variable mapping in this table:",
        "  'Proportion of area flooded' = Seasonal_Ratio (contemporaneous, year t)",
        "  NDVI_{t-1} = NDVI_median_t_minus_1 (Median model) / NDVI_mean_t_minus_1 (Mean model)",
        "  NDBI_{t-1} = NDBI_median_t_minus_1 (Median model) / NDBI_mean_t_minus_1 (Mean model)",
        "",
        "Pooled sample: 3,890 obs across 778 ACs, 5 years (2015-2019), 108 district clusters.",
        "By-state samples: Bihar 1,215 (37 districts); Jharkhand 475 (22); Odisha 735 (30); WB 1,465 (19).",
    ]
    for i, line in enumerate(notes, start=1):
        ws_n.cell(row=i, column=1, value=line)
    ws_n.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
