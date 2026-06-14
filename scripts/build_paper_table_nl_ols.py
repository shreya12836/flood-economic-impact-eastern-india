"""Publication-style regression table from NL OLS results (district-clustered)."""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
POOLED = REPO_ROOT / "outputs" / "tables" / "Regression_Results_NL_OLS_Pooled.xlsx"
BYSTATE = REPO_ROOT / "outputs" / "tables" / "Regression_Results_NL_OLS_By_State.xlsx"
OUT = REPO_ROOT / "outputs" / "tables" / "Regression_Table_NL_OLS_Paper.xlsx"

SIG_TO_STARS = {"p < 0.01": "***", "p < 0.05": "**", "p < 0.10": "*", "n.s.": ""}


def stars(sig):
    return SIG_TO_STARS.get(str(sig).strip(), "")


def fmt_coef(coef, sig):
    return f"{coef:.3f}{stars(sig)}"


def fmt_se(se):
    return f"({se:.3f})"


def get_pooled(model):
    coef = pd.read_excel(POOLED, sheet_name=f"{model}_Coef")
    stats = pd.read_excel(POOLED, sheet_name=f"{model}_Stats")
    suffix = "median" if model == "Median" else "mean"
    out = {}
    r = coef[coef["Variable"] == "Seasonal_Ratio"].iloc[0]
    out["flood"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    r = coef[coef["Variable"] == f"NDVI_{suffix}_t_minus_1"].iloc[0]
    out["ndvi"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    r = coef[coef["Variable"] == f"NDBI_{suffix}_t_minus_1"].iloc[0]
    out["ndbi"] = (r["Coefficient"], r["Std_Error"], r["Significance"])
    out["obs"] = int(stats.loc[stats["Item"] == "Observations", "Value"].iloc[0])
    out["r2"] = float(stats.loc[stats["Item"] == "R-squared", "Value"].iloc[0])
    return out


def get_state(model, state):
    df = pd.read_excel(BYSTATE, sheet_name=f"{model}_All_States")
    suffix = "median" if model == "Median" else "mean"
    coef_col, se_col, sig_col = f"{state}_Coef", f"{state}_SE", f"{state}_Sig"

    def row(varname):
        r = df[df["Variable"] == varname].iloc[0]
        return (r[coef_col], r[se_col], r[sig_col])

    return {
        "flood": row("Seasonal_Ratio"),
        "ndvi": row(f"NDVI_{suffix}_t_minus_1"),
        "ndbi": row(f"NDBI_{suffix}_t_minus_1"),
        "obs": int(df.loc[df["Variable"] == "Observations", coef_col].iloc[0]),
        "r2": float(df.loc[df["Variable"] == "R-squared", coef_col].iloc[0]),
    }


def build_sheet(ws, model_name):
    pooled = get_pooled(model_name)
    bihar = get_state(model_name, "BIHAR")
    jhar = get_state(model_name, "JHARKHAND")
    orissa = get_state(model_name, "ORISSA")
    wb_state = get_state(model_name, "WB")

    cols = [pooled, bihar, jhar, orissa, wb_state]
    headers = ["All states", "Bihar", "Jharkhand", "Odisha", "West Bengal"]

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")

    ws.cell(row=1, column=1, value="Dependent variable: Growth in Night Lights (Δ log NL) — TWFE (pyfixest)")
    ws.cell(row=1, column=1).font = bold
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    header_row = 3
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = bold
        c.alignment = center
        c.border = Border(top=thick, bottom=thin)
    ws.cell(row=header_row, column=1).border = Border(top=thick, bottom=thin)

    var_labels = [
        ("flood", "Proportion of area flooded"),
        ("ndvi", "NDVI_{t-1}"),
        ("ndbi", "NDBI_{t-1}"),
    ]

    r = header_row + 1
    for key, label in var_labels:
        ws.cell(row=r, column=1, value=label).alignment = left
        for j, col_data in enumerate(cols, start=2):
            coef, se, sig = col_data[key]
            ws.cell(row=r, column=j, value=fmt_coef(coef, sig)).alignment = center
        for j, col_data in enumerate(cols, start=2):
            coef, se, sig = col_data[key]
            cell = ws.cell(row=r + 1, column=j, value=fmt_se(se))
            cell.alignment = center
            cell.font = Font(italic=True, size=10)
        r += 2

    fe_top = r
    for label in ["AC FE", "Year FE"]:
        ws.cell(row=r, column=1, value=label).alignment = left
        for j in range(2, 7):
            ws.cell(row=r, column=j, value="Yes").alignment = center
        r += 1
    for j in range(1, 7):
        ws.cell(row=fe_top, column=j).border = Border(top=thin)

    obs_row = r
    ws.cell(row=r, column=1, value="Observations").alignment = left
    for j, col_data in enumerate(cols, start=2):
        ws.cell(row=r, column=j, value=col_data["obs"]).alignment = center
    r += 1

    ws.cell(row=r, column=1, value="R² (within)").alignment = left
    for j, col_data in enumerate(cols, start=2):
        ws.cell(row=r, column=j, value=round(col_data["r2"], 3)).alignment = center
    r += 1

    for j in range(1, 7):
        ws.cell(row=obs_row, column=j).border = Border(top=thin)
        ws.cell(row=r - 1, column=j).border = Border(bottom=thick)

    ws.column_dimensions["A"].width = 32
    for j in range(2, 7):
        ws.column_dimensions[get_column_letter(j)].width = 14


wb = Workbook()
wb.remove(wb.active)
ws_med = wb.create_sheet("Median_Model")
build_sheet(ws_med, "Median")
ws_mean = wb.create_sheet("Mean_Model")
build_sheet(ws_mean, "Mean")

ws_n = wb.create_sheet("Notes")
notes = [
    "HOW TO READ THIS TABLE",
    "----------------------",
    "Each variable cell contains TWO numbers stacked vertically:",
    "    top: estimated coefficient (β̂), with significance stars",
    "    bottom: cluster-robust standard error in parentheses (italic)",
    "",
    "Significance stars:",
    "    *** p < 0.01    ** p < 0.05    * p < 0.10    (blank) n.s.",
    "",
    "Estimator: TWFE via pyfixest. AC FE = Yes, Year FE = Yes.",
    "SEs clustered by district (correlated shocks within districts).",
    "",
    "============================================================",
    "",
    "Source files:",
    "  - Regression_Results_NL_OLS_Pooled.xlsx       (All states column)",
    "  - Regression_Results_NL_OLS_By_State.xlsx     (Bihar / Jharkhand / Odisha / WB columns)",
    "",
    "Specification:",
    "  Δlog NL_it = β₁·Seasonal_Ratio_it + β₂·NDVI_{i,t-1} + β₃·NDBI_{i,t-1} + α_i + γ_t + ε_it",
    "  α_i = AC fixed effects, γ_t = year fixed effects.",
    "  Estimator: pyfixest.feols (formula API; FE absorbed, no intercept reported).",
    "  R² shown is within-R² (after FE absorption).",
    "",
    "Variable mapping:",
    "  'Proportion of area flooded' = Seasonal_Ratio (contemporaneous, year t)",
    "  NDVI_{t-1} = NDVI_median_t_minus_1 (Median model) / NDVI_mean_t_minus_1 (Mean model)",
    "  NDBI_{t-1} = NDBI_median_t_minus_1 (Median model) / NDBI_mean_t_minus_1 (Mean model)",
]
for i, line in enumerate(notes, start=1):
    ws_n.cell(row=i, column=1, value=line)
ws_n.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"Saved: {OUT}")
